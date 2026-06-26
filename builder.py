from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from vuupt_parser import RomaneioRecord


HEADERS = {
    2:  'obs',
    3:  'vuupt',
    4:  'ticado',
    5:  'Serviço - Código ',
    6:  'Serviço - Título ',
    7:  'Serviço - Tipo',
    8:  'Agendamento - Início',
    9:  'Agendamento - Fim',
    10: 'Duração',
    11: 'Serviço - Anotação',
    12: 'Destinatário - Código',
    13: 'Destinatário - Nome ',
    14: 'Destinatário - Tipo Logradouro',
    15: 'Destinatário - Logradouro',
    16: 'Destinatário - Número ',
    17: 'Destinatário - Complemento ',
    18: 'Destinatário - Bairro ',
    19: 'Destinatário - Cidade ',
    20: 'Destinatário - Estado ',
    21: 'Destinatário - País ',
    22: 'Destinatário - CEP ',
    23: 'Destinatário - Email ',
    24: 'Destinatário - Telefone ',
    25: 'Destinatário - Latitude',
    26: 'Destinatário - Longitude',
    27: 'Destinatário - Horário de atendimento - início',
    28: 'Destinatário - Horário de atendimento - fim',
    29: 'Remetente - Código',
    30: 'Remetente - Nome',
    31: 'Remetente - Tipo Logradouro',
    32: 'Remetente - Logradouro',
    33: 'Remetente - Número',
    34: 'Remetente - Complemento',
    35: 'Remetente - Bairro',
    36: 'Remetente - Cidade',
    37: 'Remetente - Estado',
    38: 'Remetente - País',
    39: 'Remetente - CEP',
    40: 'Remetente - Email',
    41: 'Remetente - Telefone',
    42: 'Remetente - Latitude',
    43: 'Remetente - Longitude',
    44: 'Remetente - Horário de atendimento - início',
    45: 'Remetente - Horário de atendimento - fim',
    46: 'Agente - Nome',
    47: 'Campo Extra 1',
    48: 'Campo Extra 2',
    49: 'Campo Extra 3',
    50: 'Dimensão 1',
    51: 'Dimensão 2',
    52: 'Dimensão 3',
    53: 'Serviço - Habilidades',
}


def load_prev_day(xlsx_bytes: bytes) -> dict:
    """Read previous day ticagem, return {romaneio_str: (obs, vuupt, ticado)}."""
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, keep_vba=False)

    ws = wb['planilha1'] if 'planilha1' in wb.sheetnames else wb.active

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # Col A may be string or numeric
        raw = row[0]
        if isinstance(raw, float):
            key = str(int(raw))
        elif isinstance(raw, int):
            key = str(raw)
        else:
            key = str(raw).strip()

        obs    = row[1] if len(row) > 1 else None
        vuupt  = row[2] if len(row) > 2 else None
        ticado = row[3] if len(row) > 3 else None

        if any(v is not None for v in (obs, vuupt, ticado)):
            mapping[key] = (obs, vuupt, ticado)

    return mapping


def load_prev_day_full(xlsx_bytes: bytes) -> list[tuple]:
    """Return all rows (col A-D) from previous day's planilha1, for dia_anterior sheet."""
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, keep_vba=False)
    ws = wb['planilha1'] if 'planilha1' in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
        rows.append(row)
    return rows


def build_xlsx(records: list[RomaneioRecord], prev_mapping: dict, prev_full_rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'planilha1'

    # Header row
    bold = Font(bold=True)
    for col_idx, header in HEADERS.items():
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for i, rec in enumerate(records, start=2):
        prev = prev_mapping.get(rec.romaneio, (None, None, None))
        obs, vuupt, ticado = prev

        ws.cell(row=i, column=1,  value=rec.romaneio)
        ws.cell(row=i, column=2,  value=obs)
        ws.cell(row=i, column=3,  value=vuupt)
        ws.cell(row=i, column=4,  value=ticado)
        ws.cell(row=i, column=5,  value=rec.col_e)
        ws.cell(row=i, column=6,  value=rec.col_f)
        ws.cell(row=i, column=7,  value=rec.col_g)
        ws.cell(row=i, column=10, value=rec.col_j)
        ws.cell(row=i, column=11, value=rec.col_k)
        ws.cell(row=i, column=12, value=rec.col_l)
        ws.cell(row=i, column=13, value=rec.col_m)
        ws.cell(row=i, column=15, value=rec.col_o)
        ws.cell(row=i, column=24, value=rec.col_x)

    # dia_anterior sheet = previous day's planilha1 cols A-D
    ws_prev = wb.create_sheet('dia anterior')
    for row_vals in prev_full_rows:
        ws_prev.append(list(row_vals))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
